import jieba.posseg as pseg
from openpyxl import load_workbook


# 数据处理
def deal_data(excel_name1, sheet_name1, data_mapping_dict1):
    # 打开已有的excel文件
    wb = load_workbook(excel_name1)
    # 根据工作簿的名称打开一个工作簿
    sheet1 = wb[sheet_name1]
    # 对于openpyxl读取excel中sheet的内容,表格的行列下标都是从1开始
    # 遍历所有行 从第二行开始处理
    for i in range(2, sheet1.max_row + 1):
        # 只遍历第一列
        for j in range(1, 2):
            # 获取到第i行第j列的值
            word_need_to_be_split = sheet1.cell(i, j).value
            # 根据jieba分词器进行分词
            split_word, translate_word = break_up_word_and_translate(word_need_to_be_split, data_mapping_dict1)
            # 将分词后的结果保存到第i行第j+2列
            sheet1.cell(i, j + 2).value = split_word
            sheet1.cell(i, j + 3).value = translate_word
    wb.save(excel_name1)


# 分词函数
def break_up_word_and_translate(word, data_mapping_dict2):
    words = pseg.cut(word)
    # 定义一个flag,用来返回分词后是否在字典中找到匹配的英文
    translate_flag = 1
    # 分词后的结果,中文-并拼接在一起
    result = ""
    # 分词后的结果,中文-映射到英文-并拼接在一起
    english_result = ""
    for w in words:
        # 这里校验,是保证一个词语分割成多个单词导致校验值返回不准确
        if translate_flag != 0:
            translate_word = data_mapping_dict2.get(str(w.word))
            if translate_word is None:
                translate_flag = 0
                english_result = '匹配失败_'
            else:
                english_result += str(translate_word) + '_'
        result += str(w.word) + '_'
    return [result[:len(result) - 1], english_result[:len(english_result) - 1]]


# 读取银行数据字典,将中英文映射关系保存到dict中
def save_bank_dict():
    # 打开已有的excel文件
    wb = load_workbook('bank_dict.xlsx')
    # 根据工作簿的名称打开一个工作簿
    sheet1 = wb['Sheet1']
    # 定义一个字典存放中英文映射
    bank_dict = {}
    # 对于openpyxl读取excel中sheet的内容,表格的行列下标都是从1开始
    # 遍历所有行
    for i in range(1, sheet1.max_row + 1):
        # 遍历第第一列
        for j in range(1, 2):
            dict_key = sheet1.cell(i, j).value
            dict_value = sheet1.cell(i, j + 1).value
            dict_value = dict_value[1:]
            bank_dict[dict_key] = dict_value
    wb.close()
    return bank_dict


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    # 设定读取的excel名称
    excel_name = 'testWord.xlsx'
    # 设定读取的sheet的名称
    sheet_name = 'Sheet1'
    # 保存银行数据字典
    data_mapping_dict = save_bank_dict()
    # 数据处理
    deal_data(excel_name, sheet_name, data_mapping_dict)
    # 待优化
    # 1. 设定读取sheet页的名称 √
    # 2. 设定遍历的中文名称的列头 √
    # 3. 设定英文字段名称的列头
    # 4. 设定列类型的中文与英文的映射关系 -- 两种方案
    #          ①：直接设定一个字段类型名称与英文固定的映射关系,但是这种方式可能会浪费一些字段空间
    #          ②：将字段中文名称后缀与具体的英文字段(带长度限定)的匹配关系,存储在excel中,读取到dict中
    # 5. 分词后转化为英文的映射后结果check--check是否有匹配不到的中文--通过设定的匹配失败去check


# See PyCharm help at https://www.jetbrains.com/help/pycharm/
